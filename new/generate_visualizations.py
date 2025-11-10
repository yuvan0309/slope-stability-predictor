#!/usr/bin/env python3
"""
Visualization module for new FoS prediction system with Ru incorporation.
- 80% training comparison for all models
- 20% testing visualization for best model only
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import json
import os
import shutil


# Set style
plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")
COLORS = ['#2ecc71', '#3498db', '#e74c3c', '#f39c12', '#9b59b6', '#16a085']


def plot_training_comparison_all_models(training_results, output_dir):
    """
    Create comprehensive comparison charts for all models on 80% training data.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    models = list(training_results.keys())
    r2_scores = [training_results[m]['r2'] for m in models]
    rmse_scores = [training_results[m]['rmse'] for m in models]
    mae_scores = [training_results[m]['mae'] for m in models]
    
    # Find best model
    best_idx = np.argmax(r2_scores)
    best_model = models[best_idx]
    
    # 1. Training Metrics Comparison Bar Chart
    fig, axes = plt.subplots(1, 3, figsize=(18, 6))
    
    # R² Score
    bars1 = axes[0].bar(range(len(models)), r2_scores, color=COLORS[:len(models)], 
                        edgecolor='black', linewidth=1.5)
    bars1[best_idx].set_color('#FFD700')  # Gold for best
    bars1[best_idx].set_edgecolor('red')
    bars1[best_idx].set_linewidth(2.5)
    axes[0].set_xlabel('Models', fontsize=12, fontweight='bold')
    axes[0].set_ylabel('R² Score', fontsize=12, fontweight='bold')
    axes[0].set_title('Training R² Comparison (80% Data)', fontsize=14, fontweight='bold')
    axes[0].set_xticks(range(len(models)))
    axes[0].set_xticklabels(models, rotation=45, ha='right')
    axes[0].grid(axis='y', alpha=0.3)
    
    # RMSE
    bars2 = axes[1].bar(range(len(models)), rmse_scores, color=COLORS[:len(models)], 
                        edgecolor='black', linewidth=1.5)
    bars2[best_idx].set_color('#FFD700')
    bars2[best_idx].set_edgecolor('red')
    bars2[best_idx].set_linewidth(2.5)
    axes[1].set_xlabel('Models', fontsize=12, fontweight='bold')
    axes[1].set_ylabel('RMSE', fontsize=12, fontweight='bold')
    axes[1].set_title('Training RMSE Comparison (80% Data)', fontsize=14, fontweight='bold')
    axes[1].set_xticks(range(len(models)))
    axes[1].set_xticklabels(models, rotation=45, ha='right')
    axes[1].grid(axis='y', alpha=0.3)
    
    # MAE
    bars3 = axes[2].bar(range(len(models)), mae_scores, color=COLORS[:len(models)], 
                        edgecolor='black', linewidth=1.5)
    bars3[best_idx].set_color('#FFD700')
    bars3[best_idx].set_edgecolor('red')
    bars3[best_idx].set_linewidth(2.5)
    axes[2].set_xlabel('Models', fontsize=12, fontweight='bold')
    axes[2].set_ylabel('MAE', fontsize=12, fontweight='bold')
    axes[2].set_title('Training MAE Comparison (80% Data)', fontsize=14, fontweight='bold')
    axes[2].set_xticks(range(len(models)))
    axes[2].set_xticklabels(models, rotation=45, ha='right')
    axes[2].grid(axis='y', alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_dir / 'training_comparison_all_models.png', dpi=300, bbox_inches='tight')
    print(f"✓ Saved: training_comparison_all_models.png")
    plt.close()
    
    # 2. Training Metrics Table
    fig, ax = plt.subplots(figsize=(12, 8))
    ax.axis('off')
    
    table_data = [['Model', 'R²', 'RMSE', 'MAE', 'Status']]
    for i, model in enumerate(models):
        status = '🏆 BEST' if i == best_idx else ''
        table_data.append([
            model,
            f"{r2_scores[i]:.4f}",
            f"{rmse_scores[i]:.4f}",
            f"{mae_scores[i]:.4f}",
            status
        ])
    
    table = ax.table(cellText=table_data, cellLoc='center', loc='center',
                    colWidths=[0.25, 0.15, 0.15, 0.15, 0.15])
    
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1, 2.5)
    
    # Style header
    for i in range(5):
        cell = table[(0, i)]
        cell.set_facecolor('#2c3e50')
        cell.set_text_props(weight='bold', color='white', fontsize=12)
        cell.set_edgecolor('black')
        cell.set_linewidth(1.5)
    
    # Style data rows
    for row in range(1, len(table_data)):
        for col in range(5):
            cell = table[(row, col)]
            if row == best_idx + 1:  # Best model row
                cell.set_facecolor('#FFD700')
                cell.set_text_props(weight='bold')
            elif row % 2 == 0:
                cell.set_facecolor('#ecf0f1')
            else:
                cell.set_facecolor('white')
            cell.set_edgecolor('black')
            cell.set_linewidth(1)
    
    plt.title('Training Results Comparison - 80% Training Data', 
             fontsize=14, fontweight='bold', pad=20)
    plt.tight_layout()
    plt.savefig(output_dir / 'training_results_table.png', dpi=300, bbox_inches='tight')
    print(f"✓ Saved: training_results_table.png")
    plt.close()


def plot_model_testing(model_name, test_results, training_results, output_dir):
    """
    Create visualizations for a model's testing on 20% test data.
    """
    output_dir = Path(output_dir)
    y_test = np.array(test_results['actual'])
    y_pred = np.array(test_results['predictions'])
    
    # 1. Actual vs Predicted scatter plot
    fig, ax = plt.subplots(figsize=(10, 10))
    
    ax.scatter(y_test, y_pred, s=100, facecolors='none', edgecolors='black', 
              linewidth=1.5, marker='o', zorder=3)
    
    # Perfect prediction line
    min_val = min(y_test.min(), y_pred.min()) - 0.2
    max_val = max(y_test.max(), y_pred.max()) + 0.2
    ax.plot([min_val, max_val], [min_val, max_val], 'k-', 
           linewidth=2, alpha=0.8, zorder=1)
    
    ax.set_xlabel('Actual FoS', fontsize=14, fontweight='bold')
    ax.set_ylabel('Predicted FoS', fontsize=14, fontweight='bold')
    ax.set_title(f'{model_name} - Test Data Comparison (20%)', 
                fontsize=15, fontweight='bold', pad=15)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.set_xlim(min_val, max_val)
    ax.set_ylim(min_val, max_val)
    ax.set_aspect('equal', adjustable='box')
    
    plt.tight_layout()
    plt.savefig(output_dir / f'{model_name.lower().replace(" ", "_")}_test_comparison.png', 
               dpi=300, bbox_inches='tight')
    print(f"✓ Saved: {model_name.lower().replace(' ', '_')}_test_comparison.png")
    plt.close()
    
    # 2. Line graph showing actual vs predicted for test data
    test_indices = np.arange(1, len(y_test) + 1)
    
    fig, ax = plt.subplots(figsize=(12, 7))
    
    ax.plot(test_indices, y_test, 'o-', color='#3498db', linewidth=2.5, 
            markersize=10, markerfacecolor='#3498db', markeredgecolor='black', 
            markeredgewidth=1, label='Actual', alpha=0.9)
    
    ax.plot(test_indices, y_pred, 's--', color='#e67e22', linewidth=2.5, 
            markersize=9, markerfacecolor='#e67e22', markeredgecolor='black', 
            markeredgewidth=1, label='Predicted', alpha=0.9)
    
    ax.set_xlabel('Test Data Samples (20%)', fontsize=14, fontweight='bold')
    ax.set_ylabel('FoS (Factor of Safety)', fontsize=14, fontweight='bold')
    ax.set_title(f'Actual and Predicted FoS by {model_name} for Testing Data', 
                fontsize=15, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='best', fontsize=12, framealpha=0.95, edgecolor='black', fancybox=True, shadow=True)
    
    plt.tight_layout()
    plt.savefig(output_dir / f'{model_name.lower().replace(" ", "_")}_test_line_graph.png', 
               dpi=300, bbox_inches='tight')
    print(f"✓ Saved: {model_name.lower().replace(' ', '_')}_test_line_graph.png")
    plt.close()
    
    # 3. Training vs Testing comparison for best model
    fig, ax = plt.subplots(figsize=(10, 6))
    
    metrics = ['R²', 'RMSE', 'MAE']
    train_values = [
        training_results[model_name]['r2'],
        training_results[model_name]['rmse'],
        training_results[model_name]['mae']
    ]
    test_values = [
        test_results['r2'],
        test_results['rmse'],
        test_results['mae']
    ]
    
    x = np.arange(len(metrics))
    width = 0.35
    
    bars1 = ax.bar(x - width/2, train_values, width, label='Training (80%)', 
                   color='#3498db', edgecolor='black', linewidth=1.5)
    bars2 = ax.bar(x + width/2, test_values, width, label='Testing (20%)', 
                   color='#e74c3c', edgecolor='black', linewidth=1.5)
    
    ax.set_xlabel('Metrics', fontsize=13, fontweight='bold')
    ax.set_ylabel('Values', fontsize=13, fontweight='bold')
    ax.set_title(f'{model_name} - Training vs Testing Performance', 
                fontsize=14, fontweight='bold', pad=15)
    ax.set_xticks(x)
    ax.set_xticklabels(metrics)
    ax.legend(fontsize=11, framealpha=0.9, edgecolor='black')
    ax.grid(axis='y', alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_dir / f'{model_name.lower().replace(" ", "_")}_train_vs_test.png', 
               dpi=300, bbox_inches='tight')
    print(f"✓ Saved: {model_name.lower().replace(' ', '_')}_train_vs_test.png")
    plt.close()


def create_all_visualizations(results_file='models/results_summary.json', 
                            output_dir='visualizations',
                            models_dir='models'):
    """
    Create all visualizations from results file and export to CSV
    
    Args:
        results_file: Path to JSON results file
        output_dir: Directory to save visualizations
        models_dir: Directory where CSV files are saved
    """
    # Load results
    with open(results_file, 'r') as f:
        results = json.load(f)
    
    os.makedirs(output_dir, exist_ok=True)
    
    print("\n📊 Generating visualizations...")
    
    # 1. Training comparison for all models
    plot_training_comparison_all_models(results['training_results'], output_dir)
    
    # 2. Testing results for both GB and XGBoost
    if results['test_results']:
        test_results = results['test_results']
        training_results = results['training_results']
        # Plot for each tested model
        for model_name, model_test_results in test_results.items():
            plot_model_testing(model_name, model_test_results, training_results, output_dir)
    
    # 3. Create combined results Excel file
    print("\n📋 Creating Excel output files...")
    create_excel_outputs(models_dir, output_dir)
    
    print("✓ All visualizations generated successfully!")


def create_excel_outputs(models_dir='models', output_dir='visualizations'):
    """
    Create comprehensive Excel file with all results
    
    Args:
        models_dir: Directory containing CSV files
        output_dir: Directory to save Excel file
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        excel_path = os.path.join(output_dir, 'all_results.xlsx')
        
        # Create Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Load training results
            training_csv = os.path.join(models_dir, 'training_results.csv')
            if os.path.exists(training_csv):
                df_training = pd.read_csv(training_csv)
                df_training.to_excel(writer, sheet_name='Training Results (All Models)', index=False)
                print(f"  ✓ Added Training Results sheet (6 models)")
            
            # Load test results
            test_csv = os.path.join(models_dir, 'test_results.csv')
            if os.path.exists(test_csv):
                df_test = pd.read_csv(test_csv)
                df_test.to_excel(writer, sheet_name='Test Results (GB & XGBoost)', index=False)
                print(f"  ✓ Added Test Results sheet (2 models)")
            
            # Load test predictions for both models
            for model_name in ['gradient_boosting', 'xgboost']:
                predictions_csv = os.path.join(models_dir, f'test_predictions_{model_name}.csv')
                if os.path.exists(predictions_csv):
                    df_predictions = pd.read_csv(predictions_csv)
                    sheet_name = f'Test Predictions ({model_name.replace("_", " ").title()})'
                    df_predictions.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  ✓ Added {sheet_name} sheet ({len(df_predictions)} samples)")
        
        # Format the Excel file
        wb = load_workbook(excel_path)
        
        # Style settings
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data cells and adjust column widths
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Format numbers to 4 decimal places
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0000'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        print(f"\n  📊 Excel file created: {excel_path}")
        
        # Also copy individual CSVs to output directory for easy access
        csv_files = ['training_results.csv', 'test_results.csv', 
                     'test_predictions_gradient_boosting.csv', 'test_predictions_xgboost.csv']
        for csv_file in csv_files:
            src = os.path.join(models_dir, csv_file)
            dst = os.path.join(output_dir, csv_file)
            if os.path.exists(src):
                shutil.copy(src, dst)
                print(f"  ✓ Copied {csv_file} to visualizations/")
        
    except Exception as e:
        print(f"  ⚠ Warning: Could not create Excel file: {e}")
        print(f"  ℹ CSV files are available in {models_dir}/")


if __name__ == "__main__":
    # Test visualization
    results_path = Path(__file__).parent / "models" / "results_summary.json"
    viz_output = Path(__file__).parent / "visualizations"
    
    if results_path.exists():
        create_all_visualizations(results_path, viz_output)
    else:
        print("❌ No results found. Run train_models.py first!")
